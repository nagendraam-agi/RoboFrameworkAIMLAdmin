from datetime import time
from xml.dom import minidom
from pathlib import Path
import csv
import openpyxl
# from robot.utils import Application
import time
from pywinauto.application import Application


def select_file(filename):
    app = Application()
    time.sleep(5)
    app.connect(title=u'Open', found_index=0)
    dialog = app.top_window()  # get active top window (Open dialog)
    if not dialog.Edit.exists():  # check if Edit field is exists
        time.sleep(1)  # if no do again in 1 second (waiting for dialog after click)
    dialog.Edit.type_keys(f"{filename}", with_spaces=True)
    time.sleep(1)
    dialog['&OpenButton'].click()  # click Open button


def readXMLAsPerNode(your_test_param):
    first_parse_XML = minidom.parse(str(Path(__file__).parent.parent) + '/XMLdata.xml')
    data = first_parse_XML.getElementsByTagName(your_test_param)[0]
    return data.firstChild.data


def read_csv_file(self, filename):
    """This creates a keyword named "Read CSV File"

        This keyword takes one argument, which is a path to a .csv file. It
        returns a list of rows, with each row being a list of the data in
        each column.
        """
    data = []
    with open(filename, 'w+') as csvfile:
        reader = csv.reader(csvfile)
        next(reader, None)
        for row in reader:
            for i in row:
                data.append(i)
    return str(data)


def open_and_read_excel_file(sheetname, column_value, file):
    filename = file
    wb_obj = openpyxl.load_workbook(filename, data_only=True)
    sheet_obj = wb_obj.get_sheet_by_name(sheetname)
    column = int(column_value)
    m_row = sheet_obj.max_row
    my_list = []  # created an empty list
    for i in range(2, m_row):  # Here I have started the loop from 2 as I want to skip the column heading value in output
        cell_obj = sheet_obj.cell(row=i, column=column)
        # cell_obj.value = ''.join([i for i in cell_obj.value if i.isalpha()])
        # replacing_words = '(,),[0-9]'
        # for j in replacing_words:
        #     cell_obj.value = cell_obj.value.replace('replacing_words', '')
        my_list.append(cell_obj.value)
    return my_list